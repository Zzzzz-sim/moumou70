import openpyxl
import requests

def read_data(filename,sheetname):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheetname]
    max_row=sheet.max_row
    list1=[]
    for i in range(2,max_row+1,1):
        dict1 = dict(
            case_id=sheet.cell(row=i,column=1).value,
            url=sheet.cell(row=i, column=5).value,
            data=sheet.cell(row=i, column=6).value,
            expected=sheet.cell(row=i, column=7).value
        )
        list1.append(dict1)
    return list1
# read_data('test_case_api.xlsx','login')
# print(read_data('test_case_api.xlsx','login'))
def api_func(url,data):
    header_reg={'X-Lemonban-Media-Type':'lemonban.v2',
    'Content-Type':'application/json'}
    reg=requests.post(url=url,json=data,headers=header_reg)
    # print(reg.json())
    return reg.json()
def write_result(filename,sheetname,row,column,final_result):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheetname]
    # max.row=sheet.max_row
    sheet.cell(row=row,column=column).value=final_result
    wb.save(filename)
def execute_func(filename,sheetname):
    cases=read_data(filename,sheetname)
     # print(cases)
    for case in cases:
        case_id=case.get('case_id')
        # print(type(case_id))
        url=case['url']
        data=case.get('data')
        data=eval(data)
        expected=case.get('expected')
        expected=eval(expected)
        expected_msg = expected.get('msg')
        real_result = api_func(url=url,data=data)
        real_msg = real_result.get('msg')
        print('预期结果为：{}'.format(expected_msg))
        print('实际结果为：{}'.format(real_msg))
        if real_msg == expected_msg:
            print('第{}条用例通过！'.format(case_id))
            final_res = 'pass'
        else:
            print('第{}条用例不通过！'.format(case_id))
            final_res = 'fail'
        print('*'*30)
        write_result(filename,sheetname, case_id+1, 8, final_res)

execute_func('test_case_api2.xlsx','register')
execute_func('test_case_api2.xlsx','login')
#
# wb=openpyxl.load_workbook('test_case_api2.xlsx')
# wb.save('test_case_api2.xlsx')
